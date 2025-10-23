from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse, HTMLResponse
from minio import Minio
from openpyxl import load_workbook
import io

router = APIRouter()
client = Minio(
    "127.0.0.1:9000",
    access_key="minioadmin",
    secret_key="minioadmin",
    secure=False
)


@router.get("/download_pdf/{filename:path}")
async def download_pdf(filename: str):
    try:
        filename = filename.strip()
        
        response = client.get_object("files", filename)
        file_data = response.read()
        response.close()
        response.release_conn()
        
        return StreamingResponse(
            io.BytesIO(file_data),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(len(file_data))
            }
        )
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"File not found: {str(e)}")


@router.get("/download_excel/{filename:path}")
async def download_excel(filename: str):
    response = client.get_object("files", filename)
    file_data = response.read()
    response.close()
    response.release_conn()
    
    return StreamingResponse(
        io.BytesIO(file_data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/files/{filename:path}")
async def get_file(filename: str):
    try:
        filename = filename.strip()
        
        response = client.get_object("files", filename)
        file_data = response.read()
        response.close()
        response.release_conn()

        if filename.endswith(".pdf"):
            return StreamingResponse(
                io.BytesIO(file_data), 
                media_type="application/pdf",
                headers={"Content-Length": str(len(file_data))}
            )

        elif filename.endswith(".xlsx"):
            wb = load_workbook(io.BytesIO(file_data))
            sheet = wb.active

            rows = []
            for row in sheet.iter_rows(values_only=True):
                cells = "".join(f"<td>{c or ''}</td>" for c in row)
                rows.append(f"<tr>{cells}</tr>")

            html = f"""
            <html>
            <head>
                <meta charset="utf-8">
                <title>{filename}</title>
                <style>
                    table {{
                        border-collapse: collapse;
                        width: 100%;
                    }}
                    td, th {{
                        border: 1px solid #ccc;
                        padding: 5px;
                        text-align: left;
                    }}
                </style>
            </head>
            <body>
                <h2>{filename}</h2>
                <table>
                    {''.join(rows)}
                </table>
            </body>
            </html>
            """
            return HTMLResponse(content=html)

        else:
            return StreamingResponse(
                io.BytesIO(file_data), 
                media_type="application/octet-stream",
                headers={"Content-Length": str(len(file_data))}
            )
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"File not found: {str(e)}")